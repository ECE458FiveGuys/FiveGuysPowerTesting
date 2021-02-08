import csv
import io

from django.http import HttpResponse

from openpyxl import Workbook
from openpyxl.writer.excel import save_virtual_workbook

from database.services.bulk_data_services.table_enums import ModelTableColumnNames, InstrumentTableColumnNames, CalibrationEventColumnNames, SheetNames
from database.services.calibration_event_services.select_calibration_events import SelectCalibrationEvents
from database.services.in_app_service import InAppService
from database.services.instrument_services.select_instruments import SelectInstruments
from database.services.model_services.select_models import SelectModels


class Export(InAppService):
    def __init__(
            self,
            user_id,
            password,
    ):
        super().__init__(user_id=user_id, password=password, admin_only=True)

    def execute(self):
        workbook = Workbook()
        self.create_calibration_sheet(workbook)
        self.create_instrument_sheet(workbook)
        self.create_model_sheet(workbook)
        response = HttpResponse(save_virtual_workbook(workbook),
                                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="calibration_events.csv"'
        return response

    def create_calibration_sheet(self, workbook):
        worksheet = workbook.create_sheet(title=SheetNames.CALIBRATION_EVENTS.value)
        calibration_events = SelectCalibrationEvents(user_id=self.user.id, password=self.user.password, order_by="date")\
            .execute()
        worksheet.append([ModelTableColumnNames.VENDOR.value,
                          ModelTableColumnNames.MODEL_NUMBER.value,
                          InstrumentTableColumnNames.SERIAL_NUMBER.value,
                          CalibrationEventColumnNames.CALIBRATION_USERNAME.value,
                          CalibrationEventColumnNames.CALIBRATION_DATE.value,
                          CalibrationEventColumnNames.CALIBRATION_COMMENT.value])
        for calibration_event in calibration_events:
            worksheet.append([calibration_event.instrument.model.vendor,
                             calibration_event.instrument.model.model_number,
                             calibration_event.instrument.serial_number,
                             calibration_event.user.username,
                             calibration_event.date.__str__(),
                             calibration_event.comment])

    def create_instrument_sheet(self, workbook):
        worksheet = workbook.create_sheet(title=SheetNames.INSTRUMENTS.value)
        instruments = SelectInstruments(user_id=self.user.id, password=self.user.password)\
            .execute()
        worksheet.append([ModelTableColumnNames.VENDOR.value,
                          ModelTableColumnNames.MODEL_NUMBER.value,
                          InstrumentTableColumnNames.SERIAL_NUMBER.value,
                          InstrumentTableColumnNames.INSTRUMENT_COMMENT.value])
        for instrument in instruments:
            worksheet.append([instrument.model.vendor,
                             instrument.model.model_number,
                             instrument.serial_number,
                             instrument.comment])

    def create_model_sheet(self, workbook):
        worksheet = workbook.create_sheet(title=SheetNames.MODELS.value)
        models = SelectModels(user_id=self.user.id, password=self.user.password)\
            .execute()
        worksheet.append([ModelTableColumnNames.VENDOR.value,
                          ModelTableColumnNames.MODEL_NUMBER.value,
                          ModelTableColumnNames.MODEL_DESCRIPTION.value,
                          ModelTableColumnNames.MODEL_COMMENT.value,
                          ModelTableColumnNames.CALIBRATION_FREQUENCY.value])
        for model in models:
            worksheet.append([model.vendor,
                             model.model_number,
                             model.description,
                             model.comment,
                             model.calibration_frequency])