import csv
import io

from django.core.exceptions import ObjectDoesNotExist
from django.http import HttpResponse

from openpyxl import Workbook, load_workbook
from openpyxl.writer.excel import save_virtual_workbook

from database.exceptions import UserError, EntryDoesNotExistException, BulkException
from database.services.bulk_data_services.table_enums import ModelTableColumnNames, InstrumentTableColumnNames, CalibrationEventColumnNames, SheetNames
from database.services.calibration_event_services.create_calibration_event import CreateCalibrationEvent
from database.services.calibration_event_services.select_calibration_events import SelectCalibrationEvents
from database.services.in_app_service import InAppService
from database.services.instrument_services.create_instrument import CreateInstrument
from database.services.instrument_services.select_instruments import SelectInstruments
from database.services.model_services.create_model import CreateModel
from database.services.model_services.select_models import SelectModels


class Import(InAppService):
    def __init__(
            self,
            user_id,
            password,
            import_file
    ):
        self.import_file = import_file
        self.errors = []
        super().__init__(user_id=user_id, password=password, admin_only=True)

    def execute(self):
        wb = load_workbook(self.import_file)
        self.build_models(wb)
        self.build_instruments(wb)
        self.build_calibration_events(wb)
        if len(self.errors) > 0:
            raise BulkException(self.errors)

    def build_models(self, workbook):
        sheet = workbook.get_sheet_by_name(SheetNames.MODELS.value)
        for column in range(1, sheet.max_column + 1):
            vendor = sheet["A{}".format(column)].value
            model_number = sheet["B{}".format(column)].value
            description = sheet["C{}".format(column)].value
            comment = sheet["D{}".format(column)].value
            calibration_frequency = sheet["E{}".format(column)].value
            try:
                CreateModel(user_id=self.user.id,
                            password=self.user.password,
                            vendor=vendor,
                            model_number=model_number,
                            description=description,
                            comment=comment,
                            calibration_frequency=calibration_frequency)\
                    .execute()
            except UserError as e:
                self.errors.add(e)

    def build_instruments(self, workbook):
        sheet = workbook.get_sheet_by_name(SheetNames.INSTRUMENTS.value)
        for column in range(1, sheet.max_column + 1):
            vendor = sheet["A{}".format(column)].value
            model_number = sheet["B{}".format(column)].value
            serial_number = sheet["C{}".format(column)].value
            comment = sheet["D{}".format(column)].value
            try:
                model = SelectModels(user_id=self.user.id,
                             password=self.user.password,
                             vendor=vendor, model_number=model_number)\
                    .execute()\
                    .get(vendor=vendor, model_number=model_number)
                try:
                    CreateInstrument(user_id=self.user.id,
                                     password=self.user.password,
                                     model_id=model.id,
                                     serial_number=serial_number,
                                     comment=comment)\
                        .execute()
                except UserError as e:
                     self.errors.add(e)
            except ObjectDoesNotExist:
                    self.errors.add(EntryDoesNotExistException(entry_type="model", entry_id="model number {} and vendor {}".format(model_number, vendor)))

    def build_calibration_events(self, workbook):
        sheet = workbook.get_sheet_by_name(SheetNames.CALIBRATION_EVENTS.value)
        for column in range(1, sheet.max_column + 1):
            vendor = sheet["A{}".format(column)].value
            model_number = sheet["B{}".format(column)].value
            serial_number = sheet["C{}".format(column)].value
            username = sheet["D{}".format(column)].value
            date = sheet["E{}".format(column)].value
            comment = sheet["F{}".format(column)].value
            try:
                model = SelectModels(user_id=self.user.id,
                                     password=self.user.password,
                                     vendor=vendor, model_number=model_number) \
                    .execute() \
                    .get(vendor=vendor, model_number=model_number)
                try:
                    instrument = SelectInstruments(user_id=self.user.id,
                                     password=self.user.password,
                                     model_id=model.id,
                                     serial_number=serial_number) \
                        .execute()\
                        .get(model=model, serial_number=serial_number)
                    try:
                        CreateCalibrationEvent(user_id=self.user.id,
                                     password=self.user.password,
                                     instrument_id=instrument.id,
                                     date=date,
                                     comment=comment)
                    except UserError as e:
                        self.errors.add(e)
                except ObjectDoesNotExist:
                    self.errors.add(EntryDoesNotExistException(entry_type="instrument",
                                                           entry_id="model number {} and vendor {} and serial number {}".format(
                                                               model_number, vendor, serial_number)))
            except ObjectDoesNotExist:
                self.errors.add(EntryDoesNotExistException(entry_type="model",
                                                           entry_id="model number {} and vendor {}".format(
                                                               model_number, vendor)))